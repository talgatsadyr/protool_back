import re
from decimal import Decimal, InvalidOperation

from django import forms
from django.contrib import admin, messages
from django.shortcuts import redirect, render
from django.urls import path
from openpyxl import load_workbook

from .models import Category, Product

ARTICLE_SUFFIX_RE = re.compile(r'-(\d+)$')


class PriceImportForm(forms.Form):
    file = forms.FileField(label='Файл прайса (.xlsx)')


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ('name', 'parent', 'slug')
    search_fields = ('name', 'slug')
    list_filter = ('parent',)


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'article', 'category', 'price')
    search_fields = ('name', 'article', 'external_id')
    list_filter = ('category',)
    change_list_template = 'admin/products/product/change_list.html'

    def get_urls(self):
        custom_urls = [
            path('import-prices/', self.admin_site.admin_view(self.import_prices), name='products_product_import_prices'),
        ]
        return custom_urls + super().get_urls()

    def import_prices(self, request):
        if request.method == 'POST':
            form = PriceImportForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    updated, skipped = self._apply_price_file(form.cleaned_data['file'])
                except (KeyError, IndexError, ValueError) as exc:
                    messages.error(request, f'Не удалось прочитать файл: {exc}')
                else:
                    messages.success(request, f'Цены обновлены: {updated}. Товаров без совпадения по артикулу: {skipped}.')
                return redirect('..')
        else:
            form = PriceImportForm()
        return render(request, 'admin/products/product/import_prices.html', {'form': form, 'opts': self.model._meta})

    def _apply_price_file(self, uploaded_file):
        workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
        sheet = workbook.worksheets[0]

        prices_by_article = {}
        rows = sheet.iter_rows(min_row=2, values_only=True)
        for row in rows:
            if not row or row[0] is None or row[2] is None:
                continue
            article = str(row[0]).strip()
            try:
                price = Decimal(str(row[2])).quantize(Decimal('0.01'))
            except InvalidOperation:
                continue
            prices_by_article[article] = price

        to_update = []
        skipped = 0
        for product in Product.objects.all():
            article = product.article or self._extract_article(product.external_id)
            price = prices_by_article.get(article) if article else None
            if price is None:
                skipped += 1
                continue
            product.price = price
            if not product.article:
                product.article = article
            to_update.append(product)

        Product.objects.bulk_update(to_update, ['price', 'article'], batch_size=500)
        return len(to_update), skipped

    @staticmethod
    def _extract_article(external_id):
        if not external_id:
            return None
        match = ARTICLE_SUFFIX_RE.search(external_id)
        return match.group(1) if match else None
